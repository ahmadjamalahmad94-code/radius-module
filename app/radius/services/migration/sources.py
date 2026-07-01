"""استخراج وفحص المصدر — يحوّل أيّ ملف مرفوع إلى ``SourceDataset`` موحّد.

يكتشف النوع من المحتوى (magic bytes) لا الامتداد فقط، ويتعامل دفاعيًّا مع
الترميزات والملفات الكبيرة والمُشوَّهة. الأنواع المدعومة:

  • ``sqlite``    — ملف قاعدة SQLite (‎.db/.sqlite): يُفتَح للقراءة فقط ويُفحَص
                    عبر ``sqlite_master`` + ``PRAGMA table_info``.
  • ``sql_dump``  — تفريغ SQL (MySQL/MariaDB/Postgres): يُحلّل ``CREATE TABLE``
                    لترتيب الأعمدة و``INSERT INTO`` للصفوف (تسامُح مع backticks
                    وعلامات الاقتباس واللهجات).
  • ``xlsx``      — Excel: كل ورقة = جدول.
  • ``csv``       — CSV/TSV/نص: جدول واحد، مع كشف الفاصل والترويسة.
  • ``pdf``       — PDF جدوليّ عبر pdfplumber (أفضل-جهد، يتراجع بلطف).
  • ``mikrotik``  — تصدير RouterOS ‎.rsc: ‎/ppp secret + /ip hotspot user.

كلّ القيم تُطبَّع إلى نصوص. الدوال خالصة (لا Flask/DB).
"""
from __future__ import annotations

import csv
import gzip
import io
import os
import re
import sqlite3
import tempfile

from .model import SourceDataset, SourceTable

# حدود أمان — تمنع ملفًّا خبيثًا/ضخمًا من إغراق الذاكرة.
MAX_ROWS_PER_TABLE = 500_000
MAX_CELL_LEN = 8_000
MAX_COLUMNS = 256

_ENCODING_CHAIN = ("utf-8-sig", "utf-8", "cp1256", "windows-1256", "cp1252", "latin-1")
_SQLITE_MAGIC = b"SQLite format 3\x00"


# ════════════════════════════════════════════════════════════════════
# الواجهة العامّة
# ════════════════════════════════════════════════════════════════════

def sniff_source(file_bytes: bytes, filename: str = "") -> str:
    """يكتشف نوع المصدر من المحتوى أولًا ثمّ الامتداد."""
    if not file_bytes:
        return "unknown"
    head = file_bytes[:64]
    if head.startswith(_SQLITE_MAGIC):
        return "sqlite"
    if head.startswith(b"%PDF-"):
        return "pdf"
    if head.startswith(b"PK\x03\x04"):
        if b"xl/" in file_bytes[:8192] or b"[Content_Types].xml" in file_bytes[:8192]:
            return "xlsx"
    if head.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls-legacy"

    lower = (filename or "").lower()
    if lower.endswith((".db", ".sqlite", ".sqlite3")):
        # امتداد قاعدة لكن بلا توقيع — قد يكون فارغًا أو تالفًا.
        return "sqlite"
    if lower.endswith(".rsc"):
        return "mikrotik"
    if lower.endswith(".sql"):
        return "sql_dump"
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".pdf"):
        return "pdf"
    if lower.endswith((".csv", ".tsv", ".txt")):
        return "csv"

    # استدلال على المحتوى النصّي.
    sample = _decode_text(file_bytes[:65536])
    upper = sample.upper()
    if "CREATE TABLE" in upper or re.search(r"\bINSERT\s+INTO\b", upper):
        return "sql_dump"
    if re.search(r"/(ppp\s+secret|ip\s+hotspot\s+user|ip\s+hotspot\s+/user)", sample) \
            or re.search(r"^\s*add\s+name=", sample, re.MULTILINE):
        return "mikrotik"
    if any(d in sample for d in (",", ";", "\t", "|")) and "\n" in sample:
        return "csv"
    return "unknown"


def introspect_path(path: str, filename: str = "") -> SourceDataset:
    """فحص ملفّ على القرص — يدعم gzip والتدفّق للملفّات الكبيرة (تفريغ SQL
    بحجم مئات الميغابايت دون تحميلها كاملةً في الذاكرة). للأنواع الصغيرة
    (Excel/CSV/PDF/MikroTik/SQLite) يقرأ المحتوى ثمّ يفوّض لـ:func:`introspect`.
    """
    fn = filename or os.path.basename(path)
    try:
        with open(path, "rb") as fh:
            head = fh.read(8)
    except OSError as exc:
        ds = SourceDataset(fmt="unknown")
        ds.warnings.append(f"تعذّر فتح الملف: {exc}")
        return ds
    is_gz = head[:2] == b"\x1f\x8b"
    if is_gz and fn.lower().endswith(".gz"):
        fn = fn[:-3]

    def _opener():
        return gzip.open(path, "rb") if is_gz else open(path, "rb")

    with _opener() as fh:
        peek = fh.read(65536)
    fmt = sniff_source(peek, fn)

    if fmt == "sql_dump":
        ds = SourceDataset(fmt=fmt)
        try:
            with _opener() as fh:
                stream = io.TextIOWrapper(fh, encoding="utf-8", errors="replace")
                _consume_sql_statements(_iter_sql_statements(_read_chunks(stream)), ds)
        except Exception as exc:  # noqa: BLE001
            ds.warnings.append(f"خطأ أثناء قراءة تفريغ SQL: {exc}")
        ds.tables = [t for t in ds.tables if t.columns or t.rows]
        return ds

    if fmt == "sqlite" and not is_gz:
        ds = SourceDataset(fmt=fmt)
        _from_sqlite_path(path, ds)
        ds.tables = [t for t in ds.tables if t.columns or t.rows]
        return ds

    # أنواع صغيرة — اقرأ المحتوى (مفكوكًا إن كان gz) وفوّض.
    with _opener() as fh:
        data = fh.read()
    return introspect(data, fn)


def introspect(file_bytes: bytes, filename: str = "") -> SourceDataset:
    """نقطة الدخول — يكتشف النوع ويُرجع مجموعة بيانات موحّدة (للقراءة فقط)."""
    # gz في الذاكرة → فكّ ثمّ أعِد الفحص.
    if file_bytes[:2] == b"\x1f\x8b":
        try:
            file_bytes = gzip.decompress(file_bytes)
            if filename.lower().endswith(".gz"):
                filename = filename[:-3]
        except OSError:
            pass
    fmt = sniff_source(file_bytes, filename)
    ds = SourceDataset(fmt=fmt)
    try:
        if fmt == "sqlite":
            _from_sqlite(file_bytes, ds)
        elif fmt == "sql_dump":
            _from_sql_dump(file_bytes, ds)
        elif fmt == "xlsx":
            _from_xlsx(file_bytes, ds)
        elif fmt == "csv":
            _from_csv(file_bytes, ds, filename)
        elif fmt == "pdf":
            _from_pdf(file_bytes, ds)
        elif fmt == "mikrotik":
            _from_mikrotik(file_bytes, ds)
        elif fmt == "xls-legacy":
            ds.warnings.append(
                "صيغة .xls القديمة غير مدعومة — احفظ الملف بصيغة .xlsx ثم أعد الرفع.")
        else:
            ds.warnings.append(
                "تعذّر التعرّف على نوع الملف. المدعوم: قاعدة SQLite، تفريغ SQL، "
                "Excel/CSV، PDF جدوليّ، تصدير MikroTik (.rsc).")
    except Exception as exc:  # noqa: BLE001 — لا نُسقط التحليل على عطل مُستخرِج
        ds.warnings.append(f"خطأ أثناء قراءة المصدر: {exc}")
    # تنظيف: أسقط الجداول الفارغة تمامًا (بلا أعمدة وبلا صفوف).
    ds.tables = [t for t in ds.tables if t.columns or t.rows]
    return ds


# ════════════════════════════════════════════════════════════════════
# أدوات مشتركة
# ════════════════════════════════════════════════════════════════════

def _decode_text(file_bytes: bytes) -> str:
    for enc in _ENCODING_CHAIN:
        try:
            return file_bytes.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return file_bytes.decode("utf-8", errors="replace")


def _clip(value) -> str:
    if value is None:
        return ""
    s = str(value)
    if len(s) > MAX_CELL_LEN:
        s = s[:MAX_CELL_LEN]
    return s


def _looks_like_header(cells: list[str]) -> bool:
    """صفّ ترويسة: خلايا غير فارغة، أغلبها غير رقميّ، ومتمايزة."""
    cleaned = [c.strip() for c in cells]
    if not cleaned or not all(cleaned):
        return False
    if len(set(c.lower() for c in cleaned)) != len(cleaned):
        return False
    numeric = sum(1 for c in cleaned if re.fullmatch(r"[-+]?\d+(\.\d+)?", c))
    return numeric <= len(cleaned) * 0.4


def _make_columns(width: int, header: list[str] | None) -> list[str]:
    if header:
        cols, seen = [], {}
        for i in range(width):
            raw = (header[i].strip() if i < len(header) and header[i] else "") or f"col{i+1}"
            key = raw
            if key in seen:
                seen[key] += 1
                key = f"{raw}_{seen[raw]}"
            else:
                seen[key] = 0
            cols.append(key)
        return cols
    return [f"col{i+1}" for i in range(width)]


def _rows_to_table(name: str, grid: list[list[str]], origin: str,
                   *, note: str = "") -> SourceTable:
    """يحوّل شبكة خلايا (مع كشف الترويسة) إلى ``SourceTable``."""
    grid = [[_clip(c) for c in row] for row in grid if any(str(c).strip() for c in row)]
    if not grid:
        return SourceTable(name=name, origin=origin, note=note)
    width = min(max(len(r) for r in grid), MAX_COLUMNS)
    header = grid[0] if _looks_like_header(grid[0]) else None
    columns = _make_columns(width, header)
    data_rows = grid[1:] if header else grid
    rows: list[dict[str, str]] = []
    for r in data_rows[:MAX_ROWS_PER_TABLE]:
        rows.append({columns[i]: (r[i] if i < len(r) else "") for i in range(width)})
    if len(data_rows) > MAX_ROWS_PER_TABLE:
        note = (note + " ").strip() + f"(اقتُصرت إلى {MAX_ROWS_PER_TABLE} صفًّا)"
    return SourceTable(name=name, columns=columns, rows=rows, origin=origin, note=note)


# ════════════════════════════════════════════════════════════════════
# SQLite
# ════════════════════════════════════════════════════════════════════

def _from_sqlite(file_bytes: bytes, ds: SourceDataset) -> None:
    # نكتب البايتات لملف مؤقّت ونفتحه للقراءة فقط (immutable) — أسلم من فتح
    # مُدخَل غير موثوق على القرص الحيّ.
    tmp = tempfile.NamedTemporaryFile(prefix="mig_src_", suffix=".db", delete=False)
    try:
        tmp.write(file_bytes)
        tmp.flush()
        tmp.close()
        _from_sqlite_path(tmp.name, ds)
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


def _from_sqlite_path(path: str, ds: SourceDataset) -> None:
    """يفتح ملفّ SQLite على القرص للقراءة فقط (immutable) ويفحصه."""
    try:
        uri = f"file:{path}?mode=ro&immutable=1"
        conn = sqlite3.connect(uri, uri=True)
        conn.row_factory = sqlite3.Row
    except sqlite3.Error as exc:
        ds.warnings.append(f"ملف ليس قاعدة SQLite صالحة: {exc}")
        return
    try:
        names = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type IN ('table','view') "
            "AND name NOT LIKE 'sqlite_%' ORDER BY name").fetchall()]
        if not names:
            ds.warnings.append("قاعدة SQLite لا تحتوي جداول قابلة للقراءة.")
        for tname in names:
            try:
                cols = [r["name"] for r in conn.execute(
                    f'PRAGMA table_info("{tname}")').fetchall()][:MAX_COLUMNS]
                if not cols:
                    cur = conn.execute(f'SELECT * FROM "{tname}" LIMIT 1')
                    cols = [d[0] for d in (cur.description or [])][:MAX_COLUMNS]
                cur = conn.execute(
                    f'SELECT * FROM "{tname}" LIMIT {MAX_ROWS_PER_TABLE}')
                rows = []
                for raw in cur.fetchall():
                    rows.append({c: _clip(raw[c]) for c in cols if c in raw.keys()})
                ds.tables.append(SourceTable(
                    name=tname, columns=cols, rows=rows, origin="sqlite"))
            except sqlite3.Error as exc:
                ds.warnings.append(f"تعذّرت قراءة الجدول «{tname}»: {exc}")
    except sqlite3.Error as exc:
        ds.warnings.append(f"تعذّر فحص قاعدة SQLite: {exc}")
    finally:
        conn.close()


# ════════════════════════════════════════════════════════════════════
# تفريغ SQL (MySQL / MariaDB / Postgres)
# ════════════════════════════════════════════════════════════════════

def _strip_ident(tok: str) -> str:
    tok = tok.strip()
    if len(tok) >= 2 and tok[0] in "`\"[" and tok[-1] in "`\"]":
        tok = tok[1:-1]
    # schema.table → table
    if "." in tok:
        tok = tok.split(".")[-1].strip("`\"[]")
    return tok.strip()


_CREATE_RE = re.compile(
    r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?([`"\[]?[\w\.]+[`"\]]?)\s*\((.*?)\)\s*'
    r'(?:ENGINE|DEFAULT|;|AUTO_INCREMENT|COMMENT|/\*|WITHOUT|STRICT)',
    re.IGNORECASE | re.DOTALL)

_CREATE_HEAD_RE = re.compile(
    r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?([`"\[]?[\w\.]+[`"\]]?)\s*\(',
    re.IGNORECASE)


def _extract_create(stmt: str):
    """يستخرج (اسم الجدول، جسم الأعمدة) بمطابقة أقواس متوازنة — لا regex
    غير جشِع يَقطع عند أوّل «)» داخل ``int(11)``/``varchar(255)``/``enum(...)``
    (كان ذلك يُعمِّم أعمدة الجداول التي لا تُدرِج INSERT أعمدتها). يحترم
    النصوص والهروب."""
    m = _CREATE_HEAD_RE.search(stmt)
    if not m:
        return None, None
    tname = _strip_ident(m.group(1))
    i = stmt.find("(", m.end() - 1)
    if i < 0:
        return tname, None
    n = len(stmt)
    depth = 0
    j = i
    in_str = False
    q = ""
    while j < n:
        ch = stmt[j]
        if in_str:
            if ch == "\\" and q != "`":
                j += 2
                continue
            if ch == q:
                in_str = False
            j += 1
            continue
        if ch in ("'", '"', "`"):
            in_str = True
            q = ch
        elif ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth == 0:
                return tname, stmt[i + 1:j]
        j += 1
    return tname, stmt[i + 1:]      # غير متوازن — احتياط.


def _parse_create_columns(body: str) -> list[str]:
    """يستخرج أسماء الأعمدة (بالترتيب) من جسم ``CREATE TABLE`` المتسامح."""
    cols: list[str] = []
    depth = 0
    cur = []
    parts: list[str] = []
    for ch in body:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "," and depth == 0:
            parts.append("".join(cur))
            cur = []
        else:
            cur.append(ch)
    if cur:
        parts.append("".join(cur))
    _NON_COL = ("primary", "unique", "key", "constraint", "foreign",
                "index", "check", "fulltext", "spatial")
    for part in parts:
        p = part.strip()
        if not p:
            continue
        first = p.split(None, 1)[0]
        if first.lower().strip("`\"[]") in _NON_COL:
            continue
        cols.append(_strip_ident(first))
        if len(cols) >= MAX_COLUMNS:
            break
    return cols


def _split_sql_values(segment: str) -> list[list[str]]:
    """يحلّل ``(...),(...),(...)`` إلى صفوف قيم (نصوص؛ NULL→'')."""
    rows: list[list[str]] = []
    i, n = 0, len(segment)
    while i < n and len(rows) < MAX_ROWS_PER_TABLE:
        while i < n and segment[i] in " \t\r\n,":
            i += 1
        if i >= n or segment[i] != "(":
            if i < n and segment[i] == ";":
                break
            i += 1
            continue
        i += 1  # تخطّى '('
        values: list[str] = []
        cur: list[str] = []
        in_str = False
        quote = ""
        while i < n:
            ch = segment[i]
            if in_str:
                if ch == "\\" and i + 1 < n:        # هروب على نمط MySQL
                    cur.append(segment[i + 1])
                    i += 2
                    continue
                if ch == quote:
                    if i + 1 < n and segment[i + 1] == quote:  # '' داخل النصّ
                        cur.append(quote)
                        i += 2
                        continue
                    in_str = False
                    i += 1
                    continue
                cur.append(ch)
                i += 1
                continue
            if ch in ("'", '"'):
                in_str = True
                quote = ch
                i += 1
                continue
            if ch == ",":
                values.append("".join(cur).strip())
                cur = []
                i += 1
                continue
            if ch == ")":
                values.append("".join(cur).strip())
                i += 1
                break
            cur.append(ch)
            i += 1
        norm = []
        for v in values:
            vs = v.strip()
            if vs.upper() == "NULL":
                vs = ""
            norm.append(_clip(vs))
        if norm:
            rows.append(norm)
    return rows


def _from_sql_dump(file_bytes: bytes, ds: SourceDataset) -> None:
    """مسار في-الذاكرة (ملفّات صغيرة) — يمرّر النصّ كقطعة واحدة للمُستهلِك
    المتدفّق نفسه المستعمَل للملفّات الكبيرة."""
    text = _decode_text(file_bytes)
    _consume_sql_statements(_iter_sql_statements([text]), ds)


def _read_chunks(text_stream, size: int = 1 << 20):
    """يقرأ تيّار نصّ على دفعات (1MB) — تدفّق منخفض الذاكرة."""
    while True:
        chunk = text_stream.read(size)
        if not chunk:
            break
        yield chunk


def _iter_sql_statements(chunks):
    """يقسم تيّار SQL إلى عبارات كاملة — تدفّق سطريّ منخفض الذاكرة.

    يعتمد على خاصّيّة mysqldump: لا يُخرِج سطرًا جديدًا خامًا داخل قيمة نصّيّة
    قطّ (يُهرّبه ‎\\n)، فكلّ سطر جديد حدّ بنيويّ. لذا نجمع الأسطر حتى ينتهي
    سطر بـ«؛»، فتكون العبارة مكتملة — مستقلّ عن نمط الهروب (backslash أو ''‎)
    الذي يَكسر المسح الحرفيّ. سريع (تقطيع أسطر على مستوى C)."""
    buf: list[str] = []
    pending_line = ""
    for chunk in chunks:
        data = pending_line + chunk
        lines = data.split("\n")
        pending_line = lines.pop()          # آخر سطر قد يكون ناقصًا.
        for line in lines:
            buf.append(line)
            if line.rstrip().endswith(";"):
                stmt = "\n".join(buf).strip()
                buf = []
                if stmt:
                    yield stmt
        # حدّ أمان: عبارة ضخمة بلا «؛» (نادر) — لا تكدّس بلا حدّ.
        if sum(len(x) for x in buf) > 96 << 20:
            stmt = "\n".join(buf).strip()
            buf = []
            if stmt:
                yield stmt
    if pending_line:
        buf.append(pending_line)
    tail = "\n".join(buf).strip()
    if tail:
        yield tail


_INSERT_HEAD_RE = re.compile(
    r'^\s*INSERT\s+(?:IGNORE\s+)?INTO\s+([`"\[]?[\w\.]+[`"\]]?)\s*(\([^)]*\))?\s*VALUES',
    re.IGNORECASE | re.DOTALL)

_LEADING_COMMENT_RE = re.compile(r'^\s*(?:--[^\n]*\n|#[^\n]*\n|/\*.*?\*/)', re.DOTALL)


def _strip_leading_sql_comments(stmt: str) -> str:
    """يزيل تعليقات/فراغات البداية كي يبدأ النصّ بـCREATE/INSERT فعليًّا."""
    prev = None
    s = stmt
    while s != prev:
        prev = s
        s = _LEADING_COMMENT_RE.sub("", s, count=1)
    return s.strip()


def _consume_sql_statements(stmt_iter, ds: SourceDataset) -> None:
    """يبني الجداول تدرّجيًّا من تيّار عبارات SQL (CREATE/INSERT)."""
    create_cols: dict[str, list[str]] = {}
    tables: dict[str, SourceTable] = {}
    order: list[str] = []
    capped: set[str] = set()

    def _ensure_table(key: str, cols: list[str]) -> SourceTable:
        if key not in tables:
            t = SourceTable(name=key, columns=list(cols), rows=[], origin="sql_dump")
            tables[key] = t
            order.append(key)
        return tables[key]

    for stmt in stmt_iter:
        stmt = _strip_leading_sql_comments(stmt)
        s = stmt
        head = s[:12].upper()
        if head.startswith("CREATE"):
            tname, body = _extract_create(stmt)
            if tname and body:
                cols = _parse_create_columns(body)
                if cols:
                    create_cols[tname.lower()] = cols
                    # أنشئ الجدول (بنية) حتى لو لم تأتِ صفوف.
                    _ensure_table(tname.lower(), cols)
            continue
        if not head.startswith("INSERT"):
            continue
        m = _INSERT_HEAD_RE.match(stmt)
        if not m:
            continue
        key = _strip_ident(m.group(1)).lower()
        if key in capped:
            continue
        explicit_cols = None
        if m.group(2):
            explicit_cols = [_strip_ident(c) for c in m.group(2)[1:-1].split(",")]
        cols = explicit_cols or create_cols.get(key) or []
        t = _ensure_table(key, cols)
        if not t.columns and cols:
            t.columns = list(cols)
        segment = stmt[m.end():]
        rows = _split_sql_values(segment)
        for r in rows:
            if len(t.rows) >= MAX_ROWS_PER_TABLE:
                capped.add(key)
                t.note = f"(اقتُصرت إلى {MAX_ROWS_PER_TABLE} صفًّا)"
                break
            # وسّع الأعمدة إن لزم.
            if len(r) > len(t.columns):
                t.columns = list(t.columns) + [
                    f"col{i+1}" for i in range(len(t.columns), len(r))]
            t.rows.append({(t.columns[i] if i < len(t.columns) else f"col{i+1}"):
                           (r[i] if i < len(r) else "") for i in range(len(r))})

    for key in order:
        t = tables[key]
        # طبّع صفوفًا ناقصة الأعمدة.
        if t.columns:
            t.rows = [{c: row.get(c, "") for c in t.columns} for row in t.rows]
        ds.tables.append(t)

    if not any(t.rows for t in ds.tables):
        ds.warnings.append(
            "لم يُعثَر على صفوف INSERT قابلة للقراءة (عُرِضت بنية الجداول فقط).")


# ════════════════════════════════════════════════════════════════════
# Excel / CSV / PDF
# ════════════════════════════════════════════════════════════════════

def _from_xlsx(file_bytes: bytes, ds: SourceDataset) -> None:
    # المسار 1 — openpyxl (data_only). بعض الملفّات المُصدَّرة من لوحات
    # الطرف الثالث تحمل أنماطًا مشوّهة تُفشل openpyxl («expected Fill»).
    if _xlsx_via_openpyxl(file_bytes, ds):
        return
    # المسار 2 — قراءة XML الخام من الحزمة (zip) دون آليّة الأنماط.
    # مسار صلب: أيّ xlsx صالح ككائن zip يُقرأ حتى لو كسر النمطُ openpyxl.
    grids = _xlsx_via_raw_xml(file_bytes)
    if grids:
        for name, grid in grids:
            t = _rows_to_table(name, grid, "xlsx")
            if t.columns or t.rows:
                ds.tables.append(t)
    if not ds.tables:
        ds.warnings.append(
            "تعذّرت قراءة أوراق Excel (حتى عبر المسار البديل). "
            "احفظ الملف بصيغة CSV وأعد الرفع.")


def _xlsx_via_openpyxl(file_bytes: bytes, ds: SourceDataset) -> bool:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False
    for ro in (True, False):               # جرّب read_only ثمّ الكامل.
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=ro, data_only=True)
        except Exception:  # noqa: BLE001 — ننتقل للمسار الخام
            continue
        try:
            added = False
            for ws in wb.worksheets:
                grid: list[list[str]] = []
                for raw in ws.iter_rows(values_only=True):
                    grid.append([_xlsx_cell(c) for c in raw])
                    if len(grid) > MAX_ROWS_PER_TABLE + 1:
                        break
                t = _rows_to_table(ws.title, grid, "xlsx")
                if t.columns or t.rows:
                    ds.tables.append(t)
                    added = True
            return added
        except Exception:  # noqa: BLE001
            return False
        finally:
            try:
                wb.close()
            except Exception:  # noqa: BLE001
                pass
    return False


def _col_ref_to_index(ref: str) -> int:
    """‘AB12’ → فهرس العمود 0-based (27)."""
    letters = "".join(ch for ch in ref if ch.isalpha()).upper()
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - 64)
    return idx - 1 if idx > 0 else 0


def _xlsx_via_raw_xml(file_bytes: bytes) -> list[tuple[str, list[list[str]]]]:
    """يقرأ xlsx كـzip ويحلّل XML مباشرة (sharedStrings + كل sheet) دون
    openpyxl — يتجاوز أعطال الأنماط. يُعيد [(sheet_name, grid)]."""
    import xml.etree.ElementTree as ET
    import zipfile

    def local(tag: str) -> str:
        return tag.rsplit("}", 1)[-1]

    out: list[tuple[str, list[list[str]]]] = []
    try:
        zf = zipfile.ZipFile(io.BytesIO(file_bytes))
    except Exception:  # noqa: BLE001
        return out
    try:
        names = set(zf.namelist())
        # 1) الجُمل المشتركة (shared strings).
        shared: list[str] = []
        if "xl/sharedStrings.xml" in names:
            try:
                root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for si in root:
                    if local(si.tag) != "si":
                        continue
                    shared.append("".join(
                        (t.text or "") for t in si.iter() if local(t.tag) == "t"))
            except ET.ParseError:
                shared = []
        # 2) أسماء الأوراق + ترتيبها (workbook.xml + rels).
        sheet_files = _xlsx_sheet_paths(zf, names, local)
        # 3) حلّل كل ورقة.
        for sheet_name, path in sheet_files:
            if path not in names:
                continue
            grid = _xlsx_parse_sheet(zf.read(path), shared, local)
            if grid:
                out.append((sheet_name, grid))
    except Exception:  # noqa: BLE001
        return out
    finally:
        zf.close()
    return out


def _xlsx_sheet_paths(zf, names, local) -> list[tuple[str, str]]:
    import xml.etree.ElementTree as ET
    # اربط r:id → target عبر workbook.xml.rels.
    rid_target: dict[str, str] = {}
    if "xl/_rels/workbook.xml.rels" in names:
        try:
            rroot = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            for rel in rroot:
                rid = rel.attrib.get("Id", "")
                tgt = rel.attrib.get("Target", "")
                if not (rid and tgt):
                    continue
                if tgt.startswith("/"):
                    path = tgt.lstrip("/")          # مطلق من جذر الحزمة.
                elif tgt.startswith("xl/"):
                    path = tgt
                else:
                    path = "xl/" + tgt              # نسبيّ لمجلّد xl/.
                rid_target[rid] = path
        except ET.ParseError:
            pass
    sheets: list[tuple[str, str]] = []
    if "xl/workbook.xml" in names:
        try:
            wroot = ET.fromstring(zf.read("xl/workbook.xml"))
            for el in wroot.iter():
                if local(el.tag) != "sheet":
                    continue
                nm = el.attrib.get("name", f"sheet{len(sheets)+1}")
                rid = ""
                for k, v in el.attrib.items():
                    if local(k) == "id":
                        rid = v
                        break
                path = rid_target.get(rid, "")
                sheets.append((nm, path))
        except ET.ParseError:
            pass
    if not sheets:   # احتياط: خمّن أسماء الملفّات القياسيّة.
        i = 1
        while f"xl/worksheets/sheet{i}.xml" in names:
            sheets.append((f"sheet{i}", f"xl/worksheets/sheet{i}.xml"))
            i += 1
    return sheets


def _xlsx_parse_sheet(data: bytes, shared: list[str], local) -> list[list[str]]:
    import xml.etree.ElementTree as ET
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return []
    grid: list[list[str]] = []
    for row_el in root.iter():
        if local(row_el.tag) != "row":
            continue
        cells: dict[int, str] = {}
        max_c = -1
        for c in row_el:
            if local(c.tag) != "c":
                continue
            ref = c.attrib.get("r", "")
            ci = _col_ref_to_index(ref) if ref else (max_c + 1)
            ctype = c.attrib.get("t", "")
            val = ""
            if ctype == "inlineStr":
                val = "".join((t.text or "") for t in c.iter() if local(t.tag) == "t")
            else:
                vtext = ""
                for ch in c:
                    if local(ch.tag) == "v":
                        vtext = ch.text or ""
                        break
                if ctype == "s":            # فهرس جملة مشتركة.
                    try:
                        val = shared[int(vtext)]
                    except (ValueError, IndexError):
                        val = ""
                else:
                    val = vtext
            cells[ci] = _clip(val)
            max_c = max(max_c, ci)
        if not cells:
            grid.append([])
            continue
        grid.append([cells.get(i, "") for i in range(max_c + 1)])
        if len(grid) > MAX_ROWS_PER_TABLE + 1:
            break
    return grid


def _xlsx_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _from_csv(file_bytes: bytes, ds: SourceDataset, filename: str) -> None:
    text = _decode_text(file_bytes).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        ds.warnings.append("الملف فارغ.")
        return
    sample = "\n".join(text.split("\n")[:20])
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = _best_delimiter(sample)
    grid: list[list[str]] = []
    for raw in csv.reader(io.StringIO(text), delimiter=delimiter):
        grid.append([_clip(c) for c in raw])
    name = os.path.splitext(os.path.basename(filename or "data"))[0] or "data"
    t = _rows_to_table(name, grid, "csv")
    if t.columns or t.rows:
        ds.tables.append(t)
    else:
        ds.warnings.append("تعذّرت قراءة صفوف من ملف CSV.")


def _best_delimiter(sample: str) -> str:
    best = (",", -1.0)
    for d in (",", ";", "\t", "|"):
        try:
            counts = [len(r) for r in csv.reader(io.StringIO(sample), delimiter=d)
                      if any(c.strip() for c in r)]
        except csv.Error:
            continue
        if not counts:
            continue
        avg = sum(counts) / len(counts)
        if avg <= 1.0:
            continue
        var = sum((c - avg) ** 2 for c in counts) / len(counts)
        score = avg - var * 0.5
        if score > best[1]:
            best = (d, score)
    return best[0]


def _from_pdf(file_bytes: bytes, ds: SourceDataset) -> None:
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        ds.warnings.append("مكتبة قراءة PDF غير مثبّتة على الخادم (pdfplumber).")
        return
    n_table = 0
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for pi, page in enumerate(pdf.pages):
                try:
                    page_tables = page.extract_tables() or []
                except Exception:  # noqa: BLE001
                    page_tables = []
                for tbl in page_tables:
                    grid = [[_clip(c) for c in (row or [])] for row in (tbl or [])]
                    t = _rows_to_table(f"pdf_table_{pi+1}_{n_table+1}", grid, "pdf")
                    if t.rows:
                        ds.tables.append(t)
                        n_table += 1
    except Exception as exc:  # noqa: BLE001
        ds.warnings.append(f"تعذّر استخراج جداول PDF: {exc}")
    if not ds.tables:
        ds.warnings.append(
            "لم تُستخرَج جداول من PDF — قد لا يكون جدوليًّا. حوّله إلى Excel/CSV "
            "للحصول على نتيجة أدقّ.")


# ════════════════════════════════════════════════════════════════════
# MikroTik RouterOS export (.rsc)
# ════════════════════════════════════════════════════════════════════

_RSC_KV_RE = re.compile(r'([\w\-]+)=("(?:[^"\\]|\\.)*"|\S+)')


def _parse_rsc_kv(line: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for m in _RSC_KV_RE.finditer(line):
        key = m.group(1).strip()
        val = m.group(2).strip()
        if len(val) >= 2 and val[0] == '"' and val[-1] == '"':
            val = val[1:-1].replace('\\"', '"')
        out[key] = _clip(val)
    return out


def _from_mikrotik(file_bytes: bytes, ds: SourceDataset) -> None:
    text = _decode_text(file_bytes)
    # طيّ أسطر المتابعة (سطر ينتهي بـ '\' يُكمَل في التالي).
    text = re.sub(r"\\\s*\n", " ", text)
    current = ""   # القسم الحاليّ من سطر يبدأ بـ '/'
    buckets: dict[str, list[dict[str, str]]] = {}
    order: list[str] = []

    def bucket_for(section: str) -> str | None:
        s = section.lower()
        if "ppp" in s and "secret" in s:
            return "ppp_secrets"
        if "hotspot" in s and "user" in s and "profile" not in s:
            return "hotspot_users"
        if "ppp" in s and "profile" in s:
            return "ppp_profiles"
        if "hotspot" in s and "profile" in s:
            return "hotspot_profiles"
        return None

    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("/"):
            current = line.lstrip("/").strip()
            continue
        # سطر بيانات: قد يبدأ بـ add/set أو بـ kv مباشرة.
        verb = ""
        rest = line
        first = line.split(None, 1)
        if first and first[0].lower() in ("add", "set"):
            verb = first[0].lower()
            rest = first[1] if len(first) > 1 else ""
        if verb and verb != "add":
            continue
        kv = _parse_rsc_kv(rest)
        if not kv:
            continue
        bkey = bucket_for(current)
        if bkey is None:
            # قد يكون سطر add يحمل قرينة على القسم (مثلًا داخل /ppp secret).
            continue
        buckets.setdefault(bkey, [])
        if bkey not in order:
            order.append(bkey)
        buckets[bkey].append(kv)

    if not buckets:
        ds.warnings.append(
            "لم يُعثَر على ‎/ppp secret أو /ip hotspot user في تصدير MikroTik.")
        return

    for bkey in order:
        recs = buckets[bkey][:MAX_ROWS_PER_TABLE]
        cols: list[str] = []
        for r in recs:
            for k in r.keys():
                if k not in cols:
                    cols.append(k)
        cols = cols[:MAX_COLUMNS]
        rows = [{c: r.get(c, "") for c in cols} for r in recs]
        ds.tables.append(SourceTable(name=bkey, columns=cols, rows=rows,
                                     origin="mikrotik"))


__all__ = ["sniff_source", "introspect", "MAX_ROWS_PER_TABLE"]
