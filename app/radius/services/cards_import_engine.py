"""Cards import engine — multi-format intelligent parser.

Accepts CSV / Excel (.xlsx) / PDF and returns a list of
``(username, password)`` pairs without forcing the operator to declare
which column is which or in what order they appear.

**Version policy:** bump ENGINE_VERSION on every committed change to
detection logic. The UI surfaces this in a build-marker chip so the
operator can visually confirm the running container has the latest
code (no more "deployed?" guesswork).

Pipeline
--------
1. ``sniff_format(file_bytes, filename)`` — magic-byte + extension sniff.
2. ``extract_table(file_bytes, fmt)`` — format-specific extractor that
   normalises to a unified ``list[list[str]]`` shape (rows of cells).
3. ``detect_credentials(table)`` — header-aware first (Arabic + English
   synonyms), then falls back to scoring each column by content shape.

The engine deliberately keeps the heuristics conservative — when in
doubt, it returns warnings rather than guessing wrong.  The caller is
expected to surface those warnings so the operator can inspect the
preview before committing.

Pure functions. No Flask coupling.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Iterable


# ─── Engine version — bump on every detection-logic change. ─────────
#
# Visible to operators via a chip on the smart-import card and via
# the /preview JSON response. Lets us tell at a glance whether the
# running container actually has the latest code.
ENGINE_VERSION = "1.3"
ENGINE_BUILD_NOTE = "grid-pair + labelled-grid"


# ─── Public types ────────────────────────────────────────────────────

@dataclass
class Card:
    username: str
    password: str = ""


@dataclass
class DetectedColumns:
    """How the engine interpreted the extracted table."""
    username_index: int | None = None
    password_index: int | None = None
    header_row_present: bool = False
    strategy: str = "unknown"           # 'header' | 'shape-score' | 'single-column' | 'pair-token'
    column_scores: list[float] = field(default_factory=list)


@dataclass
class EngineResult:
    """The complete parsing outcome — what the route hands to the UI."""
    cards: list[Card] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    detected: DetectedColumns = field(default_factory=DetectedColumns)
    fmt: str = "unknown"                # 'csv' | 'xlsx' | 'pdf'
    rows_seen: int = 0
    rows_skipped: int = 0
    sheet_names: list[str] = field(default_factory=list)


# ─── Constants — synonyms in EN + AR ─────────────────────────────────

# All comparisons happen on a normalised key (lowercased, stripped,
# whitespace collapsed, punctuation removed). Keep entries normalised.
#
# Synonyms are split by signal strength. Strong synonyms are unambiguous
# ("username" → username column, full stop). Weak synonyms are matched
# only when no strong synonym is found anywhere in the header row — so
# a column called "No" doesn't shadow a column called "Username".
_USERNAME_SYNONYMS_STRONG = {
    # English
    "username", "user", "userid", "user id", "login", "loginname",
    "account", "accountid", "cardno", "cardnumber", "cardid",
    "vouchercode", "voucherid", "loginid", "subscriber",
    # Arabic
    "اسم المستخدم", "المستخدم", "مستخدم", "اليوزر", "يوزر",
    "رقم الكرت", "رقم البطاقة", "رقم القسيمة", "كود البطاقة",
}
_USERNAME_SYNONYMS_WEAK = {
    "card", "voucher", "ticket", "name", "id", "code", "account",
    "no", "number", "ref",
    "كرت", "الكرت", "بطاقة", "البطاقة", "قسيمة", "القسيمة",
    "تذكرة", "كود", "الكود", "اسم", "الاسم", "رقم", "الرقم",
    "حساب", "الحساب",
}

_PASSWORD_SYNONYMS_STRONG = {
    # English
    "password", "pass", "passwd", "pwd", "passcode", "accesscode",
    "pincode", "pinno",
    # Arabic
    "كلمة المرور", "كلمة السر", "كلمه المرور", "كلمه السر",
    "الباس", "الباسوورد", "باسوورد", "الرقم السري",
    "كود الدخول",
}
_PASSWORD_SYNONYMS_WEAK = {
    "secret", "pin", "key", "auth", "code",
    "السر", "المرور", "السري", "الرمز", "رمز", "مفتاح", "المفتاح",
}

# When two candidate cells appear on the same line in PDF text mode,
# these tokens are stripped before the value:
_INLINE_USERNAME_PREFIXES = (
    "username:", "user:", "login:", "account:", "card:", "id:",
    "اسم المستخدم:", "المستخدم:", "اليوزر:", "يوزر:", "الكرت:", "كرت:",
    "البطاقة:", "بطاقة:", "الاسم:", "اسم:",
)
_INLINE_PASSWORD_PREFIXES = (
    "password:", "pass:", "pwd:", "secret:", "pin:", "code:",
    "كلمة المرور:", "كلمة السر:", "الرمز:", "رمز:", "الباس:",
    "الرقم السري:", "السري:", "باسوورد:",
)

# Cell normalisation — characters that look like spaces but aren't:
_INVISIBLE_CHARS = (
    " ", "​", "‌", "‍", "‎", "‏",
    "﻿", "‪", "‫", "‬", "‭", "‮",
)


# ─── Public API ──────────────────────────────────────────────────────

def parse(file_bytes: bytes, filename: str) -> EngineResult:
    """High-level entry point — sniff, extract, detect."""
    fmt = sniff_format(file_bytes, filename)
    result = EngineResult(fmt=fmt)

    if fmt == "unknown":
        result.warnings.append(
            "تعذّر التعرّف على نوع الملف. الأنواع المدعومة: "
            "CSV / TSV / TXT / XLSX / PDF."
        )
        return result

    try:
        table, sheet_names, hint = extract_table(file_bytes, fmt)
    except _EngineExtractionError as exc:
        result.warnings.append(str(exc))
        return result
    except Exception as exc:  # noqa: BLE001 — unknown extractor failures
        result.warnings.append(f"خطأ غير متوقّع أثناء قراءة الملف: {exc}")
        return result

    result.sheet_names = sheet_names
    if not table:
        result.warnings.append("الملف فارغ أو لا يحتوي على صفوف قابلة للقراءة.")
        return result

    cards, detected, info = detect_credentials(table)
    # The PDF extractors may report a more specific strategy hint
    # (labelled-grid, grid-pair, pair-token). Surface it to the UI
    # instead of the downstream "header" rediscovery on the
    # materialised pair-table.
    if hint:
        detected.strategy = hint
    result.cards = cards
    result.detected = detected
    result.rows_seen = info["rows_seen"]
    result.rows_skipped = info["rows_skipped"]
    result.warnings.extend(info["warnings"])
    return result


def parse_text(text: str) -> EngineResult:
    """Parse pasted text (the existing textarea path) through the same
    detection logic. This lets the operator paste anything — CSV /
    TSV / «User: x  Pass: y» blocks / PDF-grid line dumps — and
    still get clean rows.
    """
    raw = (text or "").strip()
    if not raw:
        result = EngineResult(fmt="csv")
        result.warnings.append("النص فارغ.")
        return result

    table = _table_from_csv_text(raw.encode("utf-8"))
    result = EngineResult(fmt="csv")

    # Try the labelled-grid + PDF-grid pairs first when the table is
    # single-column-ish — these handle the patterns we see in
    # printable card sheets and labelled card sheets respectively.
    if not table or _avg_cols(table) < 2:
        lines = raw.splitlines()
        labelled = _extract_labelled_grid_pairs(lines)
        if labelled:
            result.cards = labelled
            result.detected.strategy = "labelled-grid"
            result.detected.header_row_present = False
            result.rows_seen = len(labelled)
            return result
        grid_pairs = _extract_grid_pairs(lines)
        if grid_pairs:
            result.cards = grid_pairs
            result.detected.strategy = "grid-pair"
            result.detected.header_row_present = False
            result.rows_seen = len(grid_pairs)
            return result
        pdf_pairs = _extract_inline_pairs(lines)
        if pdf_pairs and len(pdf_pairs) >= max(1, len(table) // 2):
            result.cards = pdf_pairs
            result.detected.strategy = "pair-token"
            result.detected.header_row_present = False
            result.rows_seen = len(pdf_pairs)
            return result

    if table:
        cards, detected, info = detect_credentials(table)
        result.cards = cards
        result.detected = detected
        result.rows_seen = info["rows_seen"]
        result.rows_skipped = info["rows_skipped"]
        result.warnings.extend(info["warnings"])
    else:
        result.warnings.append("تعذّرت قراءة الصفوف من النص الملصوق.")
    return result


def cards_to_csv(cards: Iterable[Card]) -> str:
    """Serialize cards as the canonical username,password CSV consumed
    by the existing cards_batches_import route. Quoting follows
    Python's default csv dialect — safe for round-tripping.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(["username", "password"])
    for c in cards:
        writer.writerow([c.username, c.password or ""])
    return buf.getvalue()


# ─── Format sniffing ─────────────────────────────────────────────────

def sniff_format(file_bytes: bytes, filename: str) -> str:
    if not file_bytes:
        return "unknown"
    head = file_bytes[:8]
    if head.startswith(b"%PDF-"):
        return "pdf"
    # ZIP signature: PK\x03\x04 — could be xlsx/xlsm/odt/...
    if head.startswith(b"PK\x03\x04"):
        # Best-effort xlsx detection — peek inside for the marker file.
        if _looks_like_xlsx(file_bytes):
            return "xlsx"
    # OLE2 — old .xls (BIFF). openpyxl can't read these; flag so the
    # caller can warn the operator gracefully.
    if head.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls-legacy"
    # Tie-break by extension for ambiguous text payloads.
    lower = (filename or "").lower()
    if lower.endswith((".csv", ".tsv", ".txt")):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".pdf"):
        return "pdf"
    # Heuristic — if the bytes decode as UTF-8 and contain printable
    # characters, treat as CSV/TXT.
    sample = file_bytes[:4096]
    try:
        sample.decode("utf-8")
        if any(b in sample for b in (b",", b";", b"\t", b"\n")):
            return "csv"
    except UnicodeDecodeError:
        pass
    return "unknown"


def _looks_like_xlsx(file_bytes: bytes) -> bool:
    # Cheap scan — xlsx archives always contain "xl/" entries.
    return b"xl/" in file_bytes[:8192] or b"[Content_Types].xml" in file_bytes[:8192]


# ─── Extractors ──────────────────────────────────────────────────────

class _EngineExtractionError(Exception):
    """Recoverable extraction error — message is operator-facing."""


def extract_table(file_bytes: bytes, fmt: str) -> tuple[list[list[str]], list[str], str | None]:
    """Returns (rows, sheet_names, strategy_hint).

    strategy_hint is set when the extractor itself already identified
    the column layout (e.g. labelled-grid or grid-pair PDFs). The
    caller uses it to override the strategy that detect_credentials
    would otherwise report from the materialised table.
    """
    if fmt == "csv":
        return _table_from_csv_text(file_bytes), [], None
    if fmt == "xlsx":
        rows, sheets = _table_from_xlsx(file_bytes)
        return rows, sheets, None
    if fmt == "pdf":
        rows, hint = _table_from_pdf(file_bytes)
        return rows, [], hint
    if fmt == "xls-legacy":
        raise _EngineExtractionError(
            "صيغة .xls القديمة غير مدعومة مباشرة. "
            "افتح الملف في Excel واحفظه بصيغة .xlsx ثم أعد الرفع."
        )
    raise _EngineExtractionError("نوع ملف غير مدعوم.")


# CSV / TSV / TXT --------------------------------------------------

_ENCODING_CHAIN = ("utf-8-sig", "utf-8", "cp1256", "windows-1256", "cp1252", "latin-1")


def _decode_text(file_bytes: bytes) -> str:
    for enc in _ENCODING_CHAIN:
        try:
            return file_bytes.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    # Last resort — replace undecodable bytes so we don't lose rows.
    return file_bytes.decode("utf-8", errors="replace")


def _table_from_csv_text(file_bytes: bytes) -> list[list[str]]:
    text = _decode_text(file_bytes)
    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return []

    # Sniff delimiter on a representative sample.
    sample = "\n".join(text.split("\n")[:20])
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Heuristic fallback — pick the delimiter that yields the most
        # consistent column count across the sample lines.
        delimiter = _pick_best_delimiter(sample)

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[str]] = []
    for raw_row in reader:
        cleaned = [_clean_cell(c) for c in raw_row]
        if any(cell for cell in cleaned):
            rows.append(cleaned)
    return rows


def _pick_best_delimiter(sample: str) -> str:
    candidates = [",", ";", "\t", "|"]
    best = (",", -1)  # (delimiter, score)
    for d in candidates:
        try:
            reader = csv.reader(io.StringIO(sample), delimiter=d)
            counts = [len(row) for row in reader if any(c.strip() for c in row)]
        except csv.Error:
            continue
        if not counts:
            continue
        # Reward delimiters where line widths are consistent AND > 1.
        avg = sum(counts) / len(counts)
        if avg <= 1.0:
            continue
        # Penalty for variance.
        variance = sum((c - avg) ** 2 for c in counts) / len(counts)
        score = avg - variance * 0.5
        if score > best[1]:
            best = (d, score)
    return best[0]


# XLSX -------------------------------------------------------------

def _table_from_xlsx(file_bytes: bytes) -> tuple[list[list[str]], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover — openpyxl is required
        raise _EngineExtractionError(
            "مكتبة قراءة Excel غير مثبّتة على الخادم (openpyxl)."
        ) from exc

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise _EngineExtractionError(
            f"تعذّرت قراءة ملف Excel — تأكّد أنه ملف XLSX صحيح. ({exc})"
        ) from exc

    rows: list[list[str]] = []
    sheet_names: list[str] = []
    try:
        for ws in wb.worksheets:
            sheet_names.append(ws.title)
            for raw in ws.iter_rows(values_only=True):
                cells = [_clean_cell(_xlsx_cell_to_str(c)) for c in raw]
                if any(cell for cell in cells):
                    rows.append(cells)
    finally:
        wb.close()
    return rows, sheet_names


def _xlsx_cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        # Excel often returns integers as floats. Render whole numbers
        # without the trailing ".0" so a card "12345" doesn't become
        # "12345.0".
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value)


# PDF --------------------------------------------------------------

def _table_from_pdf(file_bytes: bytes) -> tuple[list[list[str]], str | None]:
    """Try the richest extractor available, then fall back gracefully.

    Order: pdfplumber (tables + text) → pypdf (text only) → error.
    Returns (rows, strategy_hint).
    """
    # 1. pdfplumber — handles tabular PDFs cleanly.
    rows, hint = _pdf_via_pdfplumber(file_bytes)
    if rows:
        return rows, hint

    # 2. pypdf — text only; reconstruct cards by scanning lines.
    rows, hint = _pdf_via_pypdf(file_bytes)
    if rows:
        return rows, hint

    raise _EngineExtractionError(
        "تعذّر استخراج محتوى PDF. ركّب pdfplumber أو pypdf على الخادم، "
        "أو حوّل الملف إلى Excel/CSV."
    )


def _pdf_via_pdfplumber(file_bytes: bytes) -> tuple[list[list[str]], str | None]:
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        return [], None

    rows: list[list[str]] = []
    fallback_lines: list[str] = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                # Try table extraction first — handles grid-style PDFs.
                page_tables = []
                try:
                    page_tables = page.extract_tables() or []
                except Exception:  # noqa: BLE001 — fall through to text
                    page_tables = []
                table_added = False
                for tbl in page_tables:
                    for raw_row in tbl or []:
                        cells = [_clean_cell(c) for c in raw_row if c is not None]
                        if any(cells):
                            rows.append(cells)
                            table_added = True
                if not table_added:
                    text = page.extract_text() or ""
                    fallback_lines.extend(text.splitlines())
    except Exception:  # noqa: BLE001
        return [], None

    # If pdfplumber returned a "table" that is actually a labelled
    # grid (rows of [Label, val, Label, val, ...]) or a multi-card
    # grid (each row = one concatenated line with many credentials),
    # flatten it into lines so the credential strategies can pair
    # things up uniformly.
    if rows:
        flat = [" ".join(c for c in row if c) for row in rows]
        # Try labelled grid first — most discriminative.
        labelled = _extract_labelled_grid_pairs(flat)
        if labelled:
            return _pairs_to_table(labelled), "labelled-grid"
        # Try printable-card grid pattern next.
        if _rows_look_like_grid_lines(rows):
            fallback_lines = flat
            rows = []

    if rows:
        return rows, None

    return _pdf_text_to_table(fallback_lines)


def _pdf_via_pypdf(file_bytes: bytes) -> tuple[list[list[str]], str | None]:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore
        except ImportError:
            return [], None

    lines: list[str] = []
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        for page in reader.pages:
            text = page.extract_text() or ""
            lines.extend(text.splitlines())
    except Exception:  # noqa: BLE001
        return [], None

    return _pdf_text_to_table(lines)


def _pdf_text_to_table(lines: list[str]) -> tuple[list[list[str]], str | None]:
    """Shared post-processing for PDF text — tries every credential
    extraction strategy in order of specificity and returns the first
    that produces results, along with a strategy hint.
    """
    # ── Strategy 0: labelled grid ("Username val Username val ..."
    # then "Password val Password val ..."). Most specific — runs
    # first so it doesn't get clobbered by the looser pairers.
    labelled = _extract_labelled_grid_pairs(lines)
    if labelled:
        return _pairs_to_table(labelled), "labelled-grid"

    # ── Strategy A: printable-card grid (FUTURE-NET layout). ──
    grid_pairs = _extract_grid_pairs(lines)
    if grid_pairs:
        return _pairs_to_table(grid_pairs), "grid-pair"

    # ── Strategy B: inline "User: x / Pass: y" labelled pairs. ──
    pairs = _extract_inline_pairs(lines)
    if pairs:
        return _pairs_to_table(pairs), "pair-token"

    # Last attempt — treat each non-empty line as a single-column row.
    return [[ln] for ln in (_clean_cell(line) for line in lines) if ln], None


def _pairs_to_table(pairs: list[Card]) -> list[list[str]]:
    """Materialise extracted card pairs as a header + rows table
    that ``detect_credentials`` can re-ingest uniformly.
    """
    out: list[list[str]] = [["username", "password"]]
    out.extend([c.username, c.password or ""] for c in pairs)
    return out


def _rows_look_like_grid_lines(rows: list[list[str]]) -> bool:
    """Heuristic — pdfplumber sometimes returns each grid line as a
    single-cell row containing many space-separated credentials. If
    most rows look that way, treat them as flat lines so the grid
    pairer can process them.
    """
    if not rows:
        return False
    candidates = 0
    sampled = rows[:60]
    for row in sampled:
        if len(row) == 1 and row[0]:
            tokens = row[0].split()
            if len(tokens) >= 3 and all(_token_looks_like_credential(t) for t in tokens):
                candidates += 1
    return candidates >= max(2, len(sampled) * 0.4)


def _token_looks_like_credential(token: str) -> bool:
    if not token:
        return False
    if len(token) < 3 or len(token) > 24:
        return False
    if _contains_arabic(token):
        return False
    # Allow alphanumerics, dashes, underscores. Reject everything else
    # (Latin punctuation, currency symbols, dots, slashes, …).
    return bool(re.fullmatch(r"[A-Za-z0-9_\-]+", token))


def _extract_labelled_grid_pairs(lines: Iterable[str]) -> list[Card]:
    """Parse labelled-grid PDFs where each row carries the label
    repeated K times alongside its value:

        'Username 610069023347 Username 295068912347 Username ... '
        'Password 762778       Password 522728       Password ...'

    Pair line A's values (after stripping the label tokens) with the
    next compatible line B's values, column-by-column.

    Labels are matched against the same Arabic + English synonym sets
    the header strategy uses, so a sheet labelled in Arabic works too.
    """
    user_set = _USERNAME_SYNONYMS_STRONG | _USERNAME_SYNONYMS_WEAK
    pass_set = _PASSWORD_SYNONYMS_STRONG | _PASSWORD_SYNONYMS_WEAK

    def parse_line(raw: str) -> dict | None:
        line = _clean_cell(raw)
        if not line:
            return None
        tokens = line.split()
        if len(tokens) < 2 or len(tokens) % 2 != 0:
            return None

        def _label_key(t: str) -> str:
            return _normalise_key(t.rstrip(":：،,"))

        first_key = _label_key(tokens[0])
        if first_key in user_set:
            kind = "user"
            valid_set = user_set
        elif first_key in pass_set:
            kind = "pass"
            valid_set = pass_set
        else:
            return None

        values: list[str] = []
        for i in range(0, len(tokens), 2):
            label = _label_key(tokens[i])
            if label not in valid_set:
                return None
            value = tokens[i + 1]
            if not _token_looks_like_credential(value):
                return None
            values.append(value)
        if not values:
            return None
        return {"kind": kind, "values": values}

    classified = [parse_line(ln) for ln in lines]
    pairs: list[Card] = []
    used: set[int] = set()
    n = len(classified)
    i = 0
    while i < n:
        if i in used or classified[i] is None or classified[i]["kind"] != "user":
            i += 1
            continue
        a = classified[i]
        match_j = -1
        # Look ahead up to 6 lines for the matching password row.
        for j in range(i + 1, min(i + 7, n)):
            if j in used:
                continue
            b = classified[j]
            if b is None:
                continue
            if b["kind"] == "pass" and len(b["values"]) == len(a["values"]):
                match_j = j
                break
        if match_j < 0:
            i += 1
            continue
        b = classified[match_j]
        for u, p in zip(a["values"], b["values"]):
            pairs.append(Card(username=u, password=p))
        used.add(i)
        used.add(match_j)
        i = match_j + 1
    return pairs


def _extract_grid_pairs(lines: Iterable[str]) -> list[Card]:
    """Pair up consecutive lines that look like a row of usernames
    and a row of passwords from a printable card grid.

    Heuristic — for each line:
      * Split on whitespace.
      * Require ≥ 2 tokens, all credential-shaped, length variance ≤ 2.
      * Skip lines whose tokens are all identical (repeated brand
        like "FUTURE NET FUTURE NET ..." or "بطاقة ١٠ ساعات ...").
      * Skip lines containing Arabic text.

    Then walk the classified lines and pair each line A with the
    next compatible line B where:
      * A.n == B.n  (same column count).
      * |avg(A) - avg(B)| ≥ 2  (different shape — usernames vs passwords).
    """
    def classify(raw_line: str) -> dict | None:
        line = _clean_cell(raw_line)
        if not line or _contains_arabic(line):
            return None
        tokens = line.split()
        if len(tokens) < 2:
            return None
        # Repeated single-token lines are brand/desc noise, not data.
        if len(set(tokens)) == 1:
            return None
        if not all(_token_looks_like_credential(t) for t in tokens):
            return None
        lengths = [len(t) for t in tokens]
        if max(lengths) - min(lengths) > 2:
            return None
        avg = sum(lengths) / len(lengths)
        return {"tokens": tokens, "avg_len": avg, "n": len(tokens)}

    classified = [classify(ln) for ln in lines]
    pairs: list[Card] = []
    used: set[int] = set()
    i = 0
    n = len(classified)
    while i < n:
        if i in used or classified[i] is None:
            i += 1
            continue
        a = classified[i]
        # Look ahead up to 5 lines for a matching B (skipping over
        # brand/desc lines that classify() already dropped to None).
        match_j = -1
        for j in range(i + 1, min(i + 6, n)):
            if j in used:
                continue
            b = classified[j]
            if b is None:
                continue
            if b["n"] == a["n"] and abs(b["avg_len"] - a["avg_len"]) >= 2:
                match_j = j
                break
        if match_j < 0:
            i += 1
            continue
        b = classified[match_j]
        # Decide which line is usernames vs passwords. Convention:
        # usernames tend to be LONGER (account IDs / card numbers),
        # passwords tend to be SHORTER (PINs / short codes).
        if a["avg_len"] >= b["avg_len"]:
            user_tokens, pass_tokens = a["tokens"], b["tokens"]
        else:
            user_tokens, pass_tokens = b["tokens"], a["tokens"]
        for u, p in zip(user_tokens, pass_tokens):
            pairs.append(Card(username=u, password=p))
        used.add(i)
        used.add(match_j)
        i = match_j + 1
    return pairs


# ─── Detection ───────────────────────────────────────────────────────

def detect_credentials(table: list[list[str]]) -> tuple[list[Card], DetectedColumns, dict]:
    """The heart of the engine.

    Tries strategies in order:
      1. Header row — match Arabic/English synonyms.
      2. Shape scoring — pick the two columns that look most like
         credentials.
      3. Single column — when only one column is usable, treat each
         row as a username (no password).
    """
    warnings: list[str] = []
    info = {"rows_seen": 0, "rows_skipped": 0, "warnings": warnings}

    table = [row for row in table if any(cell for cell in row)]
    if not table:
        return [], DetectedColumns(), info

    # Normalise to a constant column count (pad short rows with "").
    max_cols = max(len(row) for row in table)
    table = [row + [""] * (max_cols - len(row)) for row in table]

    detected = DetectedColumns()
    detected.column_scores = [0.0] * max_cols

    # Strategy 1 — header row.
    cards = _try_header_strategy(table, detected, warnings)
    if cards is not None:
        info["rows_seen"] = len(table) - 1
        info["rows_skipped"] = (len(table) - 1) - len(cards)
        return cards, detected, info

    # Strategy 2 — shape scoring.
    detected.column_scores = _score_columns(table)
    cards = _try_shape_strategy(table, detected, warnings)
    if cards is not None:
        info["rows_seen"] = len(table)
        info["rows_skipped"] = len(table) - len(cards)
        return cards, detected, info

    # Strategy 3 — single column fallback (usernames only).
    cards = _try_single_column_strategy(table, detected, warnings)
    info["rows_seen"] = len(table)
    info["rows_skipped"] = len(table) - len(cards)
    return cards, detected, info


def _try_header_strategy(
    table: list[list[str]],
    detected: DetectedColumns,
    warnings: list[str],
) -> list[Card] | None:
    if not table:
        return None
    header = [_normalise_key(c) for c in table[0]]
    if not any(header):
        return None

    user_idx = _find_synonym_index(header, _USERNAME_SYNONYMS_STRONG)
    pass_idx = _find_synonym_index(header, _PASSWORD_SYNONYMS_STRONG)
    # Weak synonyms only fill in for missing columns — they never
    # override a strong match somewhere else in the header.
    if user_idx is None:
        user_idx = _find_synonym_index(
            header, _USERNAME_SYNONYMS_WEAK, exclude={pass_idx},
        )
    if pass_idx is None:
        pass_idx = _find_synonym_index(
            header, _PASSWORD_SYNONYMS_WEAK, exclude={user_idx},
        )
    if user_idx is None and pass_idx is None:
        return None
    if user_idx is None:
        warnings.append("لم يُعرَف عمود المستخدم — تم اختياره بأفضل تخمين.")
        # Pick the column that most often holds credential-shaped values
        # outside of the password column we already identified.
        scores = _score_columns(table[1:])
        if pass_idx is not None and 0 <= pass_idx < len(scores):
            scores[pass_idx] = -1
        user_idx = max(range(len(scores)), key=lambda i: scores[i]) if scores else 0

    detected.header_row_present = True
    detected.strategy = "header"
    detected.username_index = user_idx
    detected.password_index = pass_idx

    cards: list[Card] = []
    for row in table[1:]:
        username = _safe_cell(row, user_idx)
        password = _safe_cell(row, pass_idx) if pass_idx is not None else ""
        if _is_separator_row(row):
            continue
        if not username:
            continue
        if _looks_like_header_repeat(username, password):
            continue
        cards.append(Card(username=username, password=password))
    return cards


def _try_shape_strategy(
    table: list[list[str]],
    detected: DetectedColumns,
    warnings: list[str],
) -> list[Card] | None:
    scores = detected.column_scores
    if not scores or max(scores) <= 0:
        return None

    ranked = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)
    user_idx = ranked[0]
    pass_idx: int | None = ranked[1] if len(ranked) > 1 and scores[ranked[1]] > 0 else None

    # Heuristic — if the runner-up column is far weaker, drop it. This
    # protects against false-positive password columns (e.g. a "notes"
    # column with one alphanumeric note per row).
    if pass_idx is not None and scores[user_idx] > 0:
        ratio = scores[pass_idx] / max(scores[user_idx], 1e-6)
        if ratio < 0.30:
            pass_idx = None
            warnings.append(
                "عمود كلمة المرور غير واضح — تم استيراد أسماء المستخدمين فقط."
            )

    detected.strategy = "shape-score"
    detected.username_index = user_idx
    detected.password_index = pass_idx

    cards: list[Card] = []
    for row in table:
        if _is_separator_row(row):
            continue
        username = _safe_cell(row, user_idx)
        password = _safe_cell(row, pass_idx) if pass_idx is not None else ""
        if not username:
            continue
        if _looks_like_summary_row(row):
            continue
        cards.append(Card(username=username, password=password))
    return cards if cards else None


def _try_single_column_strategy(
    table: list[list[str]],
    detected: DetectedColumns,
    warnings: list[str],
) -> list[Card]:
    detected.strategy = "single-column"
    detected.username_index = 0
    detected.password_index = None
    warnings.append(
        "لم يُكتشف ترتيب أعمدة واضح — تم استخدام أوّل عمود قابل للقراءة "
        "كأسماء مستخدمين بدون كلمات مرور."
    )
    cards: list[Card] = []
    for row in table:
        if _is_separator_row(row):
            continue
        for cell in row:
            if cell:
                cards.append(Card(username=cell))
                break
    return cards


# ─── Column scoring ──────────────────────────────────────────────────

def _score_columns(rows: list[list[str]]) -> list[float]:
    if not rows:
        return []
    max_cols = max(len(row) for row in rows)
    scores = [0.0] * max_cols
    for col in range(max_cols):
        values = [row[col] if col < len(row) else "" for row in rows]
        scores[col] = _score_column(values)
    return scores


def _score_column(values: list[str]) -> float:
    """Score how strongly a column looks like a credential field.

    Returns 0 for an empty/unusable column, otherwise a positive
    floating-point score. Penalised heavily for shapes that look
    like indices, names, prices, dates, or free-text notes.
    """
    non_empty = [v for v in values if v]
    if not non_empty:
        return 0.0
    n = len(non_empty)
    score = 0.0

    # Per-cell signals.
    for v in non_empty:
        score += _score_credential_shape(v)

    # Column-level signals — these run on the full distribution.
    score += _score_column_distribution(non_empty)

    # Normalise so column length doesn't dominate (a 5000-row column
    # shouldn't automatically beat a 50-row column).
    return score / max(n, 1)


def _score_credential_shape(value: str) -> float:
    s = value.strip()
    if not s:
        return 0.0
    score = 0.0

    # Length sweet spot.
    if 3 <= len(s) <= 24:
        score += 2.0
    elif 25 <= len(s) <= 40:
        score += 0.5
    else:
        score -= 2.0

    # Inner whitespace is a strong "this isn't a credential" signal.
    if re.search(r"\s", s):
        score -= 3.0

    # Arabic letters anywhere — almost never a credential.
    if _contains_arabic(s):
        score -= 4.0

    # Email / URL / path-ish — usually metadata, not credentials.
    if re.search(r"[@/\\]", s):
        score -= 3.0

    # Currency / decimal / date markers.
    if re.search(r"[$€£،.,]", s) and re.search(r"\d", s):
        # numeric with separators — could be a price; demote softly.
        # Plain digits without separators are still good (card IDs).
        if "." in s or "," in s:
            digits = sum(c.isdigit() for c in s)
            if digits >= len(s) * 0.5:
                score -= 1.0

    if re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", s):
        score -= 5.0  # date

    # Reward alphanumeric mixes and short token shapes.
    if re.fullmatch(r"[A-Za-z0-9_-]{3,24}", s):
        score += 2.0
    if re.fullmatch(r"\d{4,12}", s):
        score += 1.5   # PIN / numeric card ID
    if re.search(r"[A-Za-z]", s) and re.search(r"\d", s):
        score += 1.0   # mixed alphanumeric — classic credential shape

    return score


def _score_column_distribution(values: list[str]) -> float:
    """Reward columns where most values share a similar length — that's
    the strongest signal of a generated credential set. Penalise
    columns that look like 1,2,3,4,5… (row indices)."""
    if not values:
        return 0.0
    score = 0.0

    # Length variance penalty/reward.
    lengths = [len(v) for v in values]
    avg_len = sum(lengths) / len(lengths)
    variance = sum((L - avg_len) ** 2 for L in lengths) / len(lengths)
    if variance < 1.0 and 4 <= avg_len <= 16:
        score += len(values) * 3.0  # strong reward for uniformity
    elif variance < 4.0:
        score += len(values) * 1.0

    # Row-index detector — sequential integers are not credentials.
    digits = [v for v in values if v.isdigit()]
    if len(digits) >= max(3, len(values) * 0.8):
        try:
            ints = sorted(int(v) for v in digits)
            consecutive = sum(1 for a, b in zip(ints, ints[1:]) if b - a == 1)
            if consecutive >= len(ints) - 2 and ints[0] <= 5:
                score -= len(values) * 4.0  # clearly an index
        except ValueError:
            pass

    # Uniqueness reward — credentials are usually distinct per row.
    unique = len(set(values))
    if unique >= len(values) * 0.9:
        score += len(values) * 1.0
    elif unique <= len(values) * 0.2:
        score -= len(values) * 2.0  # mostly repeated → likely a class label

    return score


# ─── Inline PDF "User: x  Pass: y" pair scanner ──────────────────────

_USER_TOKEN_RE = re.compile(
    r"(?:" + "|".join(re.escape(p) for p in _INLINE_USERNAME_PREFIXES) + r")\s*([^\s,;|]+)",
    re.IGNORECASE,
)
_PASS_TOKEN_RE = re.compile(
    r"(?:" + "|".join(re.escape(p) for p in _INLINE_PASSWORD_PREFIXES) + r")\s*([^\s,;|]+)",
    re.IGNORECASE,
)


def _extract_inline_pairs(lines: Iterable[str]) -> list[Card]:
    """Scan a stream of text lines for «User: x  Pass: y» patterns.

    Handles three layouts commonly seen in card-delivery PDFs:
      A. User+Pass on the same line.
      B. User on line N, Pass on line N+1 (or within a small window).
      C. Two adjacent lines that are clearly a credential pair without
         labels — e.g. ``ABC12345`` then ``98765432`` repeated.
    """
    cleaned = [_clean_cell(ln) for ln in lines]
    cleaned = [ln for ln in cleaned if ln]

    cards: list[Card] = []
    pending_user: str | None = None
    pending_user_age = 0

    for ln in cleaned:
        u = _USER_TOKEN_RE.search(ln)
        p = _PASS_TOKEN_RE.search(ln)

        if u and p:
            cards.append(Card(username=u.group(1), password=p.group(1)))
            pending_user = None
            pending_user_age = 0
            continue

        if u:
            pending_user = u.group(1)
            pending_user_age = 0
            continue

        if p and pending_user and pending_user_age <= 3:
            cards.append(Card(username=pending_user, password=p.group(1)))
            pending_user = None
            pending_user_age = 0
            continue

        if pending_user is not None:
            pending_user_age += 1
            if pending_user_age > 3:
                # Lost the pair — drop as username-only.
                cards.append(Card(username=pending_user))
                pending_user = None
                pending_user_age = 0

    if pending_user is not None:
        cards.append(Card(username=pending_user))

    return cards


# ─── Helpers ─────────────────────────────────────────────────────────

def _clean_cell(value) -> str:
    if value is None:
        return ""
    s = str(value)
    for ch in _INVISIBLE_CHARS:
        s = s.replace(ch, "")
    # Strip surrounding whitespace (incl. unicode whitespace classes).
    s = s.strip()
    # Collapse runs of internal whitespace — except we keep at most a
    # single space, because spaces are still meaningful in a label.
    s = re.sub(r"[ \t]+", " ", s)
    return s


def _normalise_key(value: str) -> str:
    s = _clean_cell(value).lower()
    s = re.sub(r"[^\w؀-ۿ ]+", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_synonym_index(
    header: list[str],
    synonyms: set[str],
    *,
    exclude: set[int | None] | None = None,
) -> int | None:
    blocked = {i for i in (exclude or set()) if i is not None}
    for i, cell in enumerate(header):
        if i in blocked or not cell:
            continue
        if cell in synonyms:
            return i
    # Try substring match for forgiving headers like "اسم المستخدم *".
    for i, cell in enumerate(header):
        if i in blocked or not cell:
            continue
        for syn in synonyms:
            if syn and (syn in cell or cell in syn):
                return i
    return None


def _safe_cell(row: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return row[index].strip()


def _is_separator_row(row: list[str]) -> bool:
    joined = "".join(row).strip()
    if not joined:
        return True
    return bool(re.fullmatch(r"[-=_*]+\s*", joined))


def _looks_like_header_repeat(username: str, password: str) -> bool:
    u = _normalise_key(username)
    p = _normalise_key(password)
    if u in _USERNAME_SYNONYMS_STRONG or u in _USERNAME_SYNONYMS_WEAK:
        return True
    if p and (p in _PASSWORD_SYNONYMS_STRONG or p in _PASSWORD_SYNONYMS_WEAK):
        return True
    return False


def _looks_like_summary_row(row: list[str]) -> bool:
    joined = " ".join(c for c in row if c).lower()
    keywords = (
        "total", "subtotal", "grand total", "sum", "count",
        "الإجمالي", "المجموع", "الاجمالي", "إجمالي",
    )
    return any(k in joined for k in keywords)


def _contains_arabic(s: str) -> bool:
    return any("؀" <= ch <= "ۿ" for ch in s)


def _avg_cols(table: list[list[str]]) -> float:
    if not table:
        return 0.0
    return sum(len(row) for row in table) / len(table)
