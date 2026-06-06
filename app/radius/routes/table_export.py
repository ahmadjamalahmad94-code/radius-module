# -*- coding: utf-8 -*-
"""تصدير الجداول الموحّد — endpoint واحد يخدم كل جداول النظام.

المسار:
    POST /admin/radius/export/table

يستقبل (form أو JSON):
    title    عنوان الجدول (يظهر في رأس PDF واسم الملف)
    columns  مصفوفة JSON بعناوين الأعمدة الظاهرة
    rows     مصفوفة JSON من مصفوفات (صفوف × أعمدة) — كل الصفوف لا الصفحة الحالية فقط
    fmt      pdf | xlsx | csv

الإخراج:
    * pdf  → مستند فاخر بهوية HobeRadius (pdf_theme: خط Cairo، تشكيل
             عربي، رأس/تذييل العلامة، جدول RTL منسّق).
    * xlsx → مصنّف openpyxl بورقة RTL، صف رأس بنفسجي بخط Cairo أبيض،
             صفوف زيبرا لافندر، وعرض أعمدة تلقائي.
    * csv  → UTF-8 مع BOM حتى يفتح صحيحًا في Excel العربي.

العميل يرسل عبر <form method=post target=_blank> بحقول مخفية (وليس
fetch) — هكذا يمرّ _csrf_token تلقائيًا عبر حاقن النماذج في
app/__init__.py ويُنزَّل الملف مباشرة من المتصفح بدون أي تعقيد Blob.
"""
from __future__ import annotations

import io
import json
import re
from datetime import datetime

from flask import Blueprint, Response, request

# حد أقصى دفاعي لحجم الجدول المرسل — يمنع إساءة الاستخدام دون أن
# يضايق أي جدول إداري حقيقي.
_MAX_ROWS = 20000
_MAX_COLS = 60

# الأحرف المسموحة في اسم الملف (عربي/لاتيني/أرقام) — الباقي يصير "-"
_FILENAME_BAD = re.compile(r"[^0-9A-Za-z؀-ۿ]+")


def _parse_payload() -> tuple[str, list[str], list[list[str]], str]:
    """قراءة الحمولة من form أو JSON وتطبيعها وتقليمها للحدود الآمنة."""
    if request.is_json:
        data = request.get_json(silent=True) or {}
        title = str(data.get("title") or "")
        columns = data.get("columns") or []
        rows = data.get("rows") or []
        fmt = str(data.get("fmt") or "")
    else:
        title = request.form.get("title") or ""
        fmt = request.form.get("fmt") or ""
        try:
            columns = json.loads(request.form.get("columns") or "[]")
        except ValueError:
            columns = []
        try:
            rows = json.loads(request.form.get("rows") or "[]")
        except ValueError:
            rows = []

    title = (title or "تصدير جدول").strip()[:120]
    fmt = (fmt or "csv").strip().lower()
    if not isinstance(columns, list):
        columns = []
    if not isinstance(rows, list):
        rows = []
    columns = [str(c or "") for c in columns][:_MAX_COLS]
    norm_rows: list[list[str]] = []
    for r in rows[:_MAX_ROWS]:
        if not isinstance(r, list):
            continue
        norm_rows.append([str(v if v is not None else "") for v in r[:_MAX_COLS]])
    return title, columns, norm_rows, fmt


def _filename(title: str, ext: str) -> str:
    """اسم ملف نظيف: العنوان + تاريخ اليوم + الامتداد (مرمّز للرأس)."""
    from urllib.parse import quote

    base = _FILENAME_BAD.sub("-", title).strip("-") or "export"
    name = f"{base}-{datetime.now():%Y-%m-%d}.{ext}"
    # filename* بترميز UTF-8 يدعم العربية في كل المتصفحات الحديثة،
    # مع fallback لاتيني بسيط للقدامى (لا نلمس التاريخ والامتداد).
    ascii_base = _FILENAME_BAD.sub("-", base.encode("ascii", "ignore").decode()).strip("-") or "export"
    fallback = f"{ascii_base}-{datetime.now():%Y-%m-%d}.{ext}"
    return f"attachment; filename={fallback}; filename*=UTF-8''{quote(name)}"


# ─── بنّاءو الصيغ الثلاث ────────────────────────────────────────────

def _build_csv(columns: list[str], rows: list[list[str]]) -> bytes:
    """CSV بسيط مع BOM حتى يتعرّف Excel على UTF-8 العربي تلقائيًا."""
    import csv

    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\r\n")
    if columns:
        writer.writerow(columns)
    writer.writerows(rows)
    return ("﻿" + out.getvalue()).encode("utf-8")


def _build_xlsx(title: str, columns: list[str], rows: list[list[str]]) -> bytes:
    """مصنّف Excel بهوية العلامة: ورقة RTL، رأس بنفسجي بخط Cairo أبيض،
    صفوف زيبرا لافندر، تجميد صف الرأس، وعرض أعمدة تلقائي."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from ..services.pdf_theme import (
        BRAND_INK, BRAND_LAVENDER, BRAND_LINE, BRAND_PRIMARY,
    )

    def _argb(hex_color: str) -> str:
        return "FF" + hex_color.lstrip("#").upper()

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "تصدير")[:31] or "تصدير"
    ws.sheet_view.rightToLeft = True  # اتجاه الورقة RTL للقراءة العربية

    thin = Side(style="thin", color=_argb(BRAND_LINE)[2:])
    border = Border(bottom=thin)
    head_font = Font(name="Cairo", bold=True, size=11, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor=_argb(BRAND_PRIMARY))
    body_font = Font(name="Cairo", size=10, color=_argb(BRAND_INK))
    zebra_fill = PatternFill("solid", fgColor=_argb(BRAND_LAVENDER))
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    if columns:
        ws.append(columns)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = center
        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A2"  # تثبيت صف الرأس عند التمرير

    for r_idx, row in enumerate(rows):
        ws.append(row)
        excel_row = ws.max_row
        for cell in ws[excel_row]:
            cell.font = body_font
            cell.alignment = center
            cell.border = border
            if r_idx % 2 == 1:  # زيبرا — صف لافندر بالتناوب
                cell.fill = zebra_fill

    # عرض أعمدة تلقائي حسب أطول قيمة (بحد أدنى/أقصى معقولين)
    n_cols = max(len(columns), max((len(r) for r in rows), default=0))
    for c in range(1, n_cols + 1):
        longest = len(columns[c - 1]) if c <= len(columns) else 0
        for row in rows:
            if c <= len(row):
                longest = max(longest, len(row[c - 1]))
        ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 4, 10), 45)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_pdf(title: str, columns: list[str], rows: list[list[str]]) -> bytes:
    """PDF فاخر عبر الثيم الموحّد — رأس/تذييل العلامة وجدول RTL منسّق."""
    from ..services.pdf_theme import (
        build_premium_pdf, empty_state, styled_table,
    )

    story: list = []
    if not rows:
        story.append(empty_state())
    else:
        # خط أصغر تلقائيًا للجداول العريضة حتى لا تنكسر الأعمدة
        font_size = 8.5 if len(columns) <= 8 else (7.5 if len(columns) <= 12 else 6.8)
        story.append(styled_table(columns, rows, font_size=font_size))

    return build_premium_pdf(
        title=title,
        subtitle=f"عدد السجلات: {len(rows)}",
        story=story,
        landscape_mode=len(columns) > 5,  # الجداول العريضة أفقيًا
        footer_note="HobeRadius • Hobe Hub",
    )


# ─── الـ endpoint ───────────────────────────────────────────────────

def export_table():
    """POST /admin/radius/export/table — يحوّل أي جدول مرسَل إلى ملف."""
    title, columns, rows, fmt = _parse_payload()

    if fmt == "pdf":
        payload = _build_pdf(title, columns, rows)
        return Response(
            payload,
            mimetype="application/pdf",
            headers={"Content-Disposition": _filename(title, "pdf")},
        )
    if fmt == "xlsx":
        payload = _build_xlsx(title, columns, rows)
        return Response(
            payload,
            mimetype=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
            headers={"Content-Disposition": _filename(title, "xlsx")},
        )
    # الافتراضي: CSV (يعمل دومًا حتى لو فشلت مكتبات أخرى)
    payload = _build_csv(columns, rows)
    return Response(
        payload,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": _filename(title, "csv")},
    )


def register_table_export_routes(bp: Blueprint) -> None:
    bp.add_url_rule(
        "/export/table", "export_table", export_table, methods=["POST"],
    )


__all__ = ["register_table_export_routes", "export_table"]
